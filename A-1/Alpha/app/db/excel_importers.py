import json
import re
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.security import utc_now_iso
from app.models.vsp import VspAirline, VspAirport, VspFrequency, VspNavaid, VspRunway, VspWaypoint


def _dms_to_decimal(value: str) -> float | None:
    if not value:
        return None
    text = str(value).strip().replace(" ", "")
    match = re.fullmatch(r"(\d{2,3})(\d{2})(\d{2}(?:\.\d+)?)?([NSEW])", text)
    if not match:
        return None
    deg = float(match.group(1))
    minutes = float(match.group(2))
    seconds = float(match.group(3) or 0)
    direction = match.group(4)
    decimal = deg + minutes / 60 + seconds / 3600
    if direction in {"S", "W"}:
        decimal *= -1
    return round(decimal, 8)


def _extract_first_number(text: str | None) -> int | None:
    if not text:
        return None
    match = re.search(r"(\d+)", str(text))
    return int(match.group(1)) if match else None


def _extract_first_decimal(text: str | None) -> float | None:
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", str(text))
    return float(match.group(1)) if match else None


def _extract_elevation_ft(text: str | None) -> int | None:
    value = _extract_first_decimal(text)
    if value is None:
        return None
    source = str(text).upper()
    if "M" in source and "FT" not in source:
        return int(round(value * 3.28084))
    return int(round(value))


def _parse_dimensions(text: str | None) -> tuple[int | None, int | None]:
    if not text:
        return None, None
    match = re.search(r"(\d+)\s*x\s*(\d+)", str(text), flags=re.IGNORECASE)
    if not match:
        return None, None
    return int(match.group(1)), int(match.group(2))


def _extract_bearing_deg(text: str | None) -> float | None:
    return _extract_first_decimal(text)


def _slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def import_airlines_excel(db: Session, workbook_path: str) -> dict:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    created = 0
    updated = 0
    now = utc_now_iso()
    seen_codes: set[str] = set()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue
        iata, icao, airline_name, callsign, country = row[:5]
        code = (iata or icao or "").strip() if isinstance(iata or icao, str) else str(iata or icao or "").strip()
        if not code or not airline_name:
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)
        airline_id = f"airline_{code.lower()}"
        existing = db.get(VspAirline, airline_id) or db.scalar(select(VspAirline).where(VspAirline.airline_code == code))
        payload = {
            "airline_code": code,
            "airline_name": str(airline_name).strip(),
            "airline_short_name": str(callsign).strip() if callsign else None,
            "country_name": str(country).strip() if country else None,
            "extra_json": json.dumps({"iata": iata, "icao": icao, "source": Path(workbook_path).name}, ensure_ascii=False),
            "updated_at": now,
        }
        if existing:
            for key, value in payload.items():
                setattr(existing, key, value)
            updated += 1
        else:
            db.add(VspAirline(airline_id=airline_id, created_at=now, **payload))
            created += 1

    db.commit()
    return {"created": created, "updated": updated}


def import_vsp_excel(db: Session, workbook_path: str) -> dict:
    wb = load_workbook(workbook_path, data_only=True)
    now = utc_now_iso()
    result = {
        "airport_upserted": False,
        "waypoints_created": 0,
        "waypoints_updated": 0,
        "runways_created": 0,
        "runways_updated": 0,
        "frequencies_created": 0,
        "frequencies_updated": 0,
        "navaids_created": 0,
        "navaids_updated": 0,
    }

    _import_airport_sheet(db, wb, workbook_path, now, result)
    _import_runway_sheet(db, wb, workbook_path, now, result)
    _import_frequency_sheet(db, wb, workbook_path, now, result)
    _import_navaid_sheet(db, wb, workbook_path, now, result)
    db.commit()
    return result


def _import_airport_sheet(db: Session, wb, workbook_path: str, now: str, result: dict) -> None:
    if "机场基础静态信息" not in wb.sheetnames:
        return
    ws = wb["机场基础静态信息"]
    lat = None
    lng = None
    elevation_ft = None
    for row in ws.iter_rows(values_only=True):
        cells = [str(cell).strip() for cell in row if cell is not None]
        if not cells:
            continue
        joined = " ".join(cells)
        if "ARP co-ordinates" in joined and len(cells) >= 3:
            coord_text = cells[2].split()
            if len(coord_text) >= 2:
                lat = _dms_to_decimal(coord_text[0])
                lng = _dms_to_decimal(coord_text[1])
        if "Elevation" in joined and len(cells) >= 3:
            elevation_ft = _extract_first_number(cells[2])

    if lat is None or lng is None:
        return

    airport = db.scalar(select(VspAirport).where(VspAirport.icao_code == "VHHH"))
    payload = {
        "icao_code": "VHHH",
        "iata_code": "HKG",
        "airport_name": "Hong Kong International Airport",
        "city_name": "Hong Kong",
        "country_name": "China",
        "lat": lat,
        "lng": lng,
        "elevation_ft": elevation_ft,
        "extra_json": json.dumps({"source": Path(workbook_path).name, "sheet": "机场基础静态信息"}, ensure_ascii=False),
        "updated_at": now,
    }
    if airport:
        for key, value in payload.items():
            setattr(airport, key, value)
    else:
        db.add(VspAirport(airport_id="airport_vhhh", created_at=now, **payload))
    db.flush()
    result["airport_upserted"] = True


def _import_navaid_sheet(db: Session, wb, workbook_path: str, now: str, result: dict) -> None:
    if "Navaid数据" not in wb.sheetnames:
        return
    ws = wb["Navaid数据"]
    current = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        aid_type, ident, frequency, hours, coord_part1, elevation, remarks = (list(row) + [None] * 7)[:7]
        if aid_type or ident or frequency or remarks:
            current = {
                "aid_type": str(aid_type).strip() if aid_type else None,
                "ident": str(ident).strip() if ident else None,
                "frequency": str(frequency).strip() if frequency else None,
                "hours": str(hours).strip() if hours else None,
                "coord_part1": str(coord_part1).strip() if coord_part1 else None,
                "coord_part2": None,
                "elevation": str(elevation).strip() if elevation else None,
                "remarks": str(remarks).strip() if remarks else None,
            }
            continue
        if current and coord_part1:
            current["coord_part2"] = str(coord_part1).strip()
            current["continuation_type"] = str(aid_type).strip() if aid_type else None
            _upsert_navaid_waypoint(db, current, workbook_path, now, result)
            current = None


def _import_runway_sheet(db: Session, wb, workbook_path: str, now: str, result: dict) -> None:
    if "跑道数据" not in wb.sheetnames:
        return
    airport = db.scalar(select(VspAirport).where(VspAirport.icao_code == "VHHH"))
    if not airport:
        return
    ws = wb["跑道数据"]
    current = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        designator, bearing, dimensions, strength_surface, threshold_coord, threshold_elev = (list(row) + [None] * 6)[:6]
        if designator:
            current = {
                "designator": str(designator).strip(),
                "true_bearing": str(bearing).strip() if bearing else None,
                "dimensions": str(dimensions).strip() if dimensions else None,
                "strength_surface": str(strength_surface).strip() if strength_surface else None,
                "threshold_lat_text": str(threshold_coord).strip() if threshold_coord else None,
                "threshold_elev_text": str(threshold_elev).strip() if threshold_elev else None,
            }
            continue
        if current and bearing:
            current["magnetic_bearing"] = str(bearing).strip()
            current["surface_type"] = str(strength_surface).strip() if strength_surface else None
            current["threshold_lng_text"] = str(threshold_coord).strip() if threshold_coord else None
            current["continuation_elev_text"] = str(threshold_elev).strip() if threshold_elev else None
            _upsert_runway(db, airport.airport_id, current, workbook_path, now, result)
            current = None


def _import_frequency_sheet(db: Session, wb, workbook_path: str, now: str, result: dict) -> None:
    if "通信频率数据" not in wb.sheetnames:
        return
    airport = db.scalar(select(VspAirport).where(VspAirport.icao_code == "VHHH"))
    if not airport:
        return
    ws = wb["通信频率数据"]
    current_service_designator = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        service_designator, callsign, frequency, hours_of_operation, remarks = (list(row) + [None] * 5)[:5]
        if service_designator:
            current_service_designator = str(service_designator).strip()
        if not frequency:
            continue
        _upsert_frequency(
            db=db,
            airport_id=airport.airport_id,
            service_designator=current_service_designator,
            callsign=str(callsign).strip() if callsign else None,
            frequency=str(frequency).strip(),
            hours_of_operation=str(hours_of_operation).strip() if hours_of_operation else None,
            remarks=str(remarks).strip() if remarks else None,
            workbook_path=workbook_path,
            now=now,
            result=result,
        )


def _upsert_navaid_waypoint(db: Session, current: dict, workbook_path: str, now: str, result: dict) -> None:
    lat = _dms_to_decimal(current.get("coord_part1"))
    lng = _dms_to_decimal(current.get("coord_part2"))
    ident = current.get("ident") or current.get("aid_type")
    if lat is None or lng is None or not ident:
        return
    airport = db.scalar(select(VspAirport).where(VspAirport.icao_code == "VHHH"))
    if not airport:
        return
    navaid_id = f"navaid_{airport.airport_id}_{_slugify(ident)}"
    navaid_extra_json = json.dumps(
        {
            "continuation_type": current.get("continuation_type"),
            "source": Path(workbook_path).name,
            "sheet": "Navaid数据",
        },
        ensure_ascii=False,
    )
    navaid = db.get(VspNavaid, navaid_id)
    navaid_payload = {
        "airport_id": airport.airport_id,
        "ident": ident,
        "name": current.get("aid_type") or ident,
        "navaid_type": current.get("aid_type"),
        "frequency": current.get("frequency"),
        "lat": lat,
        "lng": lng,
        "elevation_ft": _extract_elevation_ft(current.get("elevation")),
        "hours_of_operation": current.get("hours"),
        "remarks": current.get("remarks"),
        "extra_json": navaid_extra_json,
        "updated_at": now,
    }
    if navaid:
        for key, value in navaid_payload.items():
            setattr(navaid, key, value)
        result["navaids_updated"] += 1
    else:
        db.add(VspNavaid(navaid_id=navaid_id, created_at=now, **navaid_payload))
        result["navaids_created"] += 1

    waypoint_id = f"navaid_{ident.lower()}"
    existing = db.get(VspWaypoint, waypoint_id)
    payload = {
        "name": ident,
        "type": "navaid",
        "lat": lat,
        "lng": lng,
        "description": current.get("aid_type"),
        "extra_json": json.dumps(
            {
                "frequency": current.get("frequency"),
                "hours": current.get("hours"),
                "elevation": current.get("elevation"),
                "remarks": current.get("remarks"),
                "airport_id": airport.airport_id,
                "navaid_id": navaid_id,
                "source": Path(workbook_path).name,
                "sheet": "Navaid数据",
            },
            ensure_ascii=False,
        ),
        "updated_at": now,
    }
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        result["waypoints_updated"] += 1
    else:
        db.add(VspWaypoint(waypoint_id=waypoint_id, created_at=now, **payload))
        result["waypoints_created"] += 1


def _upsert_runway(db: Session, airport_id: str, current: dict, workbook_path: str, now: str, result: dict) -> None:
    designator = current.get("designator")
    if not designator:
        return
    threshold_lat = _dms_to_decimal(current.get("threshold_lat_text"))
    threshold_lng = _dms_to_decimal(current.get("threshold_lng_text"))
    runway_length_m, runway_width_m = _parse_dimensions(current.get("dimensions"))
    runway_id = f"runway_{airport_id}_{_slugify(designator)}"
    existing = db.get(VspRunway, runway_id)
    payload = {
        "airport_id": airport_id,
        "runway_designator": designator,
        "surface_type": current.get("surface_type"),
        "runway_length_m": runway_length_m,
        "runway_width_m": runway_width_m,
        "bearing_deg": _extract_bearing_deg(current.get("true_bearing")),
        "threshold_lat": threshold_lat,
        "threshold_lng": threshold_lng,
        "elevation_ft": _extract_elevation_ft(current.get("threshold_elev_text") or current.get("continuation_elev_text")),
        "remarks": None,
        "extra_json": json.dumps(
            {
                "true_bearing": current.get("true_bearing"),
                "magnetic_bearing": current.get("magnetic_bearing"),
                "strength_surface": current.get("strength_surface"),
                "source": Path(workbook_path).name,
                "sheet": "跑道数据",
            },
            ensure_ascii=False,
        ),
        "updated_at": now,
    }
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        result["runways_updated"] += 1
    else:
        db.add(VspRunway(runway_id=runway_id, created_at=now, **payload))
        result["runways_created"] += 1


def _upsert_frequency(
    db: Session,
    airport_id: str,
    service_designator: str | None,
    callsign: str | None,
    frequency: str,
    hours_of_operation: str | None,
    remarks: str | None,
    workbook_path: str,
    now: str,
    result: dict,
) -> None:
    frequency_key = f"{airport_id}_{service_designator or 'unknown'}_{callsign or 'unknown'}_{frequency}"
    frequency_id = f"freq_{_slugify(frequency_key)}"
    existing = db.get(VspFrequency, frequency_id)
    payload = {
        "airport_id": airport_id,
        "service_designator": service_designator,
        "callsign": callsign,
        "frequency": frequency,
        "hours_of_operation": hours_of_operation,
        "remarks": remarks,
        "extra_json": json.dumps({"source": Path(workbook_path).name, "sheet": "通信频率数据"}, ensure_ascii=False),
        "updated_at": now,
    }
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
        result["frequencies_updated"] += 1
    else:
        db.add(VspFrequency(frequency_id=frequency_id, created_at=now, **payload))
        result["frequencies_created"] += 1
